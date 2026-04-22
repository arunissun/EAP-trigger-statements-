"""
Phase 0.1c — Combination Matrix: Consolidate + Export to Excel.

Performs two tasks in sequence:

  1. CONSOLIDATION
     Merges fragmented Temperature subcategories produced by the LLM enrichment pass
     (which auto-inferred one row per unique LLM output string).  The 15 raw
     Temperature rows are collapsed into 7 canonical subcategories:

       Min temperature            Max temperature
       Cold wave                  Heatwave
       Heat index                 Wind chill index
       Temperature anomaly / departure from normal

  2. EXCEL EXPORT
     Writes combination_matrix_v1.xlsx with:
       - Sheet "Matrix"        — the full approved table, colour-coded by canonical_variable
       - Sheet "Legend"        — field definitions and allowed enum values
       - Sheet "Stats"         — row / EAP counts per canonical variable

     The XLSX is the deliverable for human review (Phase 0.1c).  Once the reviewer
     signs off, it becomes the contract for Phase 1 taxonomy matching.

Usage (from project root):
    python -m scripts.ui_normalization.matrix03_export_excel
    python -m scripts.ui_normalization.matrix03_export_excel --no-consolidate   # skip merge step
    python -m scripts.ui_normalization.matrix03_export_excel --output path/to/out.xlsx
"""

from __future__ import annotations

import argparse
import logging
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

try:
    from .config import (
        COMBINATION_MATRIX_ENRICHED_CSV,
        COMBINATION_MATRIX_XLSX,
    )
except ImportError:
    import sys
    sys.path.insert(0, str(Path(__file__).parent))
    from config import (
        COMBINATION_MATRIX_ENRICHED_CSV,
        COMBINATION_MATRIX_XLSX,
    )

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Consolidation — Temperature subcategory merge map
# ---------------------------------------------------------------------------

# Subcategories to drop entirely (their data absorbed into the target below)
_TEMPERATURE_DROP = {
    # "Cold Wave" is NOT dropped — it survives and is renamed to "Cold wave" in the loop
    "Cold wave event",        # → absorbed into "Cold wave"
    "Daytime temperature",    # → "Max temperature"
    "General temperature",    # → "Min temperature"
    "Heatwave intensity",     # → "Heatwave"
    "Heatwave occurrence",    # → "Heatwave"
    "Heatwave occurrence probability",  # → "Heatwave"
    "Max daily temp",         # → "Max temperature"
    "Max daily temperature",  # → "Max temperature"
}

# Full specification for each canonical survivor
# Keys: must match the subcategory value in the output row
_TEMPERATURE_CANONICAL: dict[str, dict] = {
    "Min temperature": {
        "primary_unit": "°C|percentile",
        "accumulation_window_applicable": "no",
        "accumulation_window_units": "n/a",
        "persistence_applicable": "yes",
        "persistence_units": "days",
        "probability_applicable": "optional",
        "lead_time_applicable": "yes",
        "lead_time_units": "days",
        "geographic_scope_required": "optional",
        "example_statement": (
            "7-day temperature forecast minimum below -40°C (-35°C in Karaganda region) with "
            "precipitation and increased wind. Regional thresholds vary: Northern Kazakhstan "
            "-40°C, Akmola -40°C, Kostanay -40°C."
        ),
        "hazard_types": "cold_wave",
        "notes": (
            "Min temperature threshold in °C or percentile below the seasonal normal. "
            "Persistence applies for cold wave definitions (typically 3+ consecutive days). "
            "Regional scope applicable when different thresholds apply per administrative area."
        ),
        "data_source": "auto_inferred",
    },
    "Max temperature": {
        "primary_unit": "°C|percentile",
        "accumulation_window_applicable": "no",
        "accumulation_window_units": "n/a",
        "persistence_applicable": "yes",
        "persistence_units": "days",
        "probability_applicable": "optional",
        "lead_time_applicable": "yes",
        "lead_time_units": "days",
        "geographic_scope_required": "optional",
        "example_statement": (
            "Expected maximum temperature higher than the 95th percentile of monthly historical "
            "maximum temperatures for 4 consecutive days. NWP forecast lead time up to 7 days."
        ),
        "hazard_types": "heatwave",
        "notes": (
            "Max temperature threshold in °C or percentile. Merges: Max daily temp, Max daily "
            "temperature, Daytime temperature. Persistence applies for heatwave definitions. "
            "Lead time from NWP model (typically 3–7 days)."
        ),
        "data_source": "auto_inferred",
    },
    "Cold wave": {
        "primary_unit": "event",
        "accumulation_window_applicable": "no",
        "accumulation_window_units": "n/a",
        "persistence_applicable": "yes",
        "persistence_units": "days",
        "probability_applicable": "optional",
        "lead_time_applicable": "optional",
        "lead_time_units": "days",
        "geographic_scope_required": "optional",
        "example_statement": (
            "72-hour forecast confirmation with at least 75% probability of cold wave occurrence "
            "for 4 consecutive days (min temp ≤ 7°C)."
        ),
        "hazard_types": "cold_wave",
        "notes": (
            "Cold wave is a binary event (occurs / does not occur). Persistence always applies — "
            "standard definition requires 3+ consecutive days below the threshold. Probability "
            "is optional when expressed as forecast confidence."
        ),
        "data_source": "auto_inferred",
        "row_count": 0,
        "source_eap_count": 0,
    },
    "Heatwave": {
        "primary_unit": "event|days",
        "accumulation_window_applicable": "no",
        "accumulation_window_units": "n/a",
        "persistence_applicable": "yes",
        "persistence_units": "days",
        "probability_applicable": "optional",
        "lead_time_applicable": "yes",
        "lead_time_units": "days",
        "geographic_scope_required": "optional",
        "example_statement": (
            "Forecast lead time maximum 5 days; heatwave occurrence confirmed when max temp "
            "exceeds 95th percentile for 3 or more consecutive days."
        ),
        "hazard_types": "heatwave",
        "notes": (
            "Heatwave is a sustained heat event (duration in days). Persistence always applies — "
            "standard definition requires 3+ consecutive days above threshold. Lead time typically "
            "3–5 days. Merges: Heatwave intensity, Heatwave occurrence, Heatwave occurrence probability."
        ),
        "data_source": "auto_inferred",
        "row_count": 0,
        "source_eap_count": 0,
    },
    "Heat index": {
        # Keep as-is from the enriched CSV; overrides below fix known issues
        "primary_unit": "°C|percentile",
        "hazard_types": "heatwave",
        "notes": (
            "Heat index combines temperature and humidity. Maps to the same data sources as "
            "temperature (ECMWF ENS, national met). Persistence applicable for sustained heat "
            "stress conditions."
        ),
    },
    "Wind chill index": {
        "primary_unit": "°C|°F",
        "hazard_types": "cold_wave",
        "notes": (
            "Wind chill index combines temperature and wind speed. Persistence applicable — "
            "cold wave conditions require sustained exposure. Same data sources as Temperature."
        ),
    },
    "Temperature anomaly / departure from normal": {
        # Keep as-is; only normalise units
        "primary_unit": "°C|°F|percent",
    },
}


def _consolidate_temperature(df: pd.DataFrame) -> pd.DataFrame:
    """
    Replace all Temperature rows with the 7 canonical subcategories defined in
    _TEMPERATURE_CANONICAL, dropping the fragmented auto-inferred duplicates.

    Row counts and source_eap_count from the dropped rows are summed and applied
    to the corresponding canonical survivor.
    """
    temp_mask = df["canonical_variable"] == "Temperature"
    temp_df = df[temp_mask].copy()
    other_df = df[~temp_mask].copy()

    # Accumulate provenance counts from rows that will be dropped
    _count_absorb: dict[str, dict[str, int]] = {
        "Min temperature": {"row_count": 0, "source_eap_count": 0},
        "Max temperature": {"row_count": 0, "source_eap_count": 0},
        "Cold wave":       {"row_count": 0, "source_eap_count": 0},
        "Heatwave":        {"row_count": 0, "source_eap_count": 0},
    }
    _absorption_map = {
        # "Cold Wave" survives (renamed); counts from "Cold wave event" absorbed
        "Cold wave event":                  "Cold wave",
        "Daytime temperature":              "Max temperature",
        "General temperature":              "Min temperature",
        "Heatwave intensity":               "Heatwave",
        "Heatwave occurrence":              "Heatwave",
        "Heatwave occurrence probability":  "Heatwave",
        "Max daily temp":                   "Max temperature",
        "Max daily temperature":            "Max temperature",
    }

    for _, row in temp_df[temp_df["subcategory"].isin(_TEMPERATURE_DROP)].iterrows():
        target = _absorption_map.get(row["subcategory"])
        if target and target in _count_absorb:
            _count_absorb[target]["row_count"] += int(row.get("row_count") or 0)
            _count_absorb[target]["source_eap_count"] += int(row.get("source_eap_count") or 0)

    # Keep survivors (rows NOT in the drop set)
    surviving = temp_df[~temp_df["subcategory"].isin(_TEMPERATURE_DROP)].copy()

    # Apply canonical overrides to each survivor
    result_rows: list[dict] = []
    for _, row in surviving.iterrows():
        sub = row["subcategory"]
        overrides = _TEMPERATURE_CANONICAL.get(sub, {})
        row_dict = row.to_dict()
        # Rename "Cold Wave" → "Cold wave" (capitalisation fix on the survivor itself)
        if sub == "Cold Wave":
            row_dict["subcategory"] = "Cold wave"
            sub = "Cold wave"
        row_dict.update(overrides)
        # Absorb dropped-row counts
        if sub in _count_absorb:
            row_dict["row_count"] = (int(row_dict.get("row_count") or 0)
                                     + _count_absorb[sub]["row_count"])
            row_dict["source_eap_count"] = (int(row_dict.get("source_eap_count") or 0)
                                             + _count_absorb[sub]["source_eap_count"])
        result_rows.append(row_dict)

    consolidated_temp = pd.DataFrame(result_rows, columns=df.columns)

    result = pd.concat([other_df, consolidated_temp], ignore_index=True)
    # Sort: canonical_variable alphabetically, then subcategory
    result = result.sort_values(
        ["canonical_variable", "subcategory"], key=lambda s: s.str.lower()
    ).reset_index(drop=True)
    return result


# ---------------------------------------------------------------------------
# Colour palette — one fill per canonical variable
# ---------------------------------------------------------------------------

_CANONICAL_COLOURS: dict[str, str] = {
    "Agricultural Impact":   "E2EFDA",  # soft green
    "Alert/Warning Status":  "FCE4D6",  # soft orange
    "Fire Risk":             "FFE699",  # soft yellow
    "Humidity":              "DDEBF7",  # soft blue
    "Hydrological Flow":     "BDD7EE",  # medium blue
    "Infectious Disease":    "F4CCCC",  # soft red
    "Other":                 "F2F2F2",  # light grey
    "Population Impact":     "EAD1DC",  # soft pink
    "Precipitation":         "C9DAF8",  # periwinkle
    "Temperature":           "FFF2CC",  # pale yellow
    "Volcanic Activity":     "D9D9D9",  # medium grey
    "Wind":                  "D0E0E3",  # teal-tinted
}

_HEADER_FILL   = PatternFill("solid", fgColor="1F497D")
_HEADER_FONT   = Font(bold=True, color="FFFFFF", size=10)
_BODY_FONT     = Font(size=9)
_BORDER_SIDE   = Side(style="thin", color="BFBFBF")
_THIN_BORDER   = Border(
    left=_BORDER_SIDE, right=_BORDER_SIDE,
    top=_BORDER_SIDE,  bottom=_BORDER_SIDE,
)

_WRAP = Alignment(wrap_text=True, vertical="top")
_CENTER = Alignment(horizontal="center", vertical="top", wrap_text=True)

# Column display widths (characters)
_COL_WIDTHS: dict[str, int] = {
    "canonical_variable":              22,
    "subcategory":                     30,
    "primary_unit":                    18,
    "accumulation_window_applicable":  14,
    "accumulation_window_units":       16,
    "persistence_applicable":          12,
    "persistence_units":               12,
    "probability_applicable":          12,
    "lead_time_applicable":            12,
    "lead_time_units":                 14,
    "geographic_scope_required":       14,
    "example_statement":               55,
    "hazard_types":                    22,
    "notes":                           45,
    "data_source":                     14,
    "row_count":                        9,
    "source_eap_count":                12,
}

# Columns where centre-alignment looks better
_CENTRE_COLS = {
    "accumulation_window_applicable",
    "accumulation_window_units",
    "persistence_applicable",
    "persistence_units",
    "probability_applicable",
    "lead_time_applicable",
    "lead_time_units",
    "geographic_scope_required",
    "data_source",
    "row_count",
    "source_eap_count",
}


def _apply_matrix_sheet(ws, df: pd.DataFrame) -> None:
    """Write the main Matrix sheet with colour-coded canonical bands."""
    cols = list(df.columns)

    # --- Header row ---
    for col_idx, col_name in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _CENTER
        cell.border = _THIN_BORDER
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    # --- Data rows ---
    for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
        canonical = str(row.get("canonical_variable", "Other"))
        hex_colour = _CANONICAL_COLOURS.get(canonical, "FFFFFF")
        row_fill = PatternFill("solid", fgColor=hex_colour)

        for col_idx, col_name in enumerate(cols, start=1):
            raw = row[col_name]
            value = "" if (raw is None or (isinstance(raw, float) and
                           __import__("math").isnan(raw))) else raw
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = row_fill
            cell.font = _BODY_FONT
            cell.border = _THIN_BORDER
            if col_name in _CENTRE_COLS:
                cell.alignment = _CENTER
            else:
                cell.alignment = _WRAP

    # --- Column widths ---
    for col_idx, col_name in enumerate(cols, start=1):
        width = _COL_WIDTHS.get(col_name, 15)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # --- Auto-filter ---
    ws.auto_filter.ref = ws.dimensions

    # --- Row heights (body rows) ---
    for row_idx in range(2, len(df) + 2):
        ws.row_dimensions[row_idx].height = 36


def _apply_legend_sheet(ws) -> None:
    """Write a Legend sheet explaining each field and its allowed values."""
    LEGEND = [
        ("Field", "Description", "Allowed Values"),
        ("canonical_variable",            "Top-level hazard variable class",
         "Agricultural Impact | Alert/Warning Status | Fire Risk | Humidity | Hydrological Flow | "
         "Infectious Disease | Other | Population Impact | Precipitation | Temperature | "
         "Volcanic Activity | Wind"),
        ("subcategory",                   "Specific type within the canonical", "Free text"),
        ("primary_unit",                  "Pipe-separated list of valid measurement units",
         "e.g. mm|cm|inches  or  °C|percentile"),
        ("accumulation_window_applicable","Does 'amount measured over X time' apply?",
         "yes | no | conditional"),
        ("accumulation_window_units",     "Valid time-window units or n/a",
         "hours|days|weeks|months  or  n/a"),
        ("persistence_applicable",        "Does 'sustained for X consecutive days/hours' apply?",
         "yes | no"),
        ("persistence_units",             "Valid persistence units or n/a", "hours|days  or  n/a"),
        ("probability_applicable",        "Can a forecast probability (%) be attached?",
         "yes | no | optional"),
        ("lead_time_applicable",          "Can a forecast lead time be attached?",
         "yes | no | optional"),
        ("lead_time_units",               "Valid lead-time units or n/a",
         "hours|days|weeks|months  or  n/a"),
        ("geographic_scope_required",     "Is geographic scope mandatory, optional, or not applicable?",
         "always | optional | not_applicable"),
        ("example_statement",             "Realistic EAP trigger statement (verbatim from corpus if possible)",
         "Free text"),
        ("hazard_types",                  "Pipe-separated hazard codes",
         "flood|drought|heatwave|cholera|cold_wave|cyclone|volcano|other"),
        ("notes",                         "Domain rule, exception, or reviewer note", "Free text"),
        ("data_source",                   "Origin of this row",
         "auto_inferred (from corpus) | llm_proposed (new row from LLM)"),
        ("row_count",                     "Number of normalized threshold rows matching this combination",
         "Integer ≥ 0"),
        ("source_eap_count",              "Number of distinct EAP documents contributing rows",
         "Integer ≥ 0"),
    ]

    hdr_fill = PatternFill("solid", fgColor="1F497D")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)

    for row_idx, row_data in enumerate(LEGEND, start=1):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = _THIN_BORDER
            cell.alignment = _WRAP
            if row_idx == 1:
                cell.fill = hdr_fill
                cell.font = hdr_font
            else:
                cell.font = _BODY_FONT

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 48
    ws.column_dimensions["C"].width = 60
    for row_idx in range(1, len(LEGEND) + 1):
        ws.row_dimensions[row_idx].height = 32


def _apply_stats_sheet(ws, df: pd.DataFrame) -> None:
    """Write a Stats sheet showing row and EAP counts per canonical variable."""
    stats = (
        df.groupby("canonical_variable")
        .agg(
            subcategory_count=("subcategory", "count"),
            total_row_count=("row_count", "sum"),
            total_eap_count=("source_eap_count", "sum"),
        )
        .reset_index()
        .sort_values("canonical_variable")
    )
    stats.columns = [
        "Canonical Variable",
        "Matrix Rows (subcategories)",
        "Corpus Rows Matched",
        "EAP Documents",
    ]

    hdr_fill = PatternFill("solid", fgColor="1F497D")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)

    for col_idx, col_name in enumerate(stats.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = _CENTER
        cell.border = _THIN_BORDER

    for row_idx, (_, row) in enumerate(stats.iterrows(), start=2):
        canonical = str(row["Canonical Variable"])
        hex_colour = _CANONICAL_COLOURS.get(canonical, "FFFFFF")
        row_fill = PatternFill("solid", fgColor=hex_colour)
        for col_idx, value in enumerate(row.tolist(), start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = row_fill
            cell.font = _BODY_FONT
            cell.border = _THIN_BORDER
            cell.alignment = _CENTER

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 20
    for row_idx in range(1, len(stats) + 2):
        ws.row_dimensions[row_idx].height = 22


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def consolidate_and_export(
    input_csv: Path = COMBINATION_MATRIX_ENRICHED_CSV,
    output_xlsx: Path = COMBINATION_MATRIX_XLSX,
    apply_consolidation: bool = True,
) -> pd.DataFrame:
    """
    Load the enriched CSV, optionally consolidate Temperature rows, and write the
    Excel file.  Returns the finalised DataFrame (also written back to input_csv).

    Args:
        input_csv:            Path to combination_matrix_enriched.csv
        output_xlsx:          Path to write combination_matrix_v1.xlsx
        apply_consolidation:  If True, merge fragmented Temperature rows first.

    Returns:
        Finalised DataFrame
    """
    if not input_csv.exists():
        raise FileNotFoundError(
            f"Enriched matrix not found at {input_csv}. "
            "Run --phase0-matrix-enrich first."
        )

    logger.info("Loading enriched matrix from %s", input_csv)
    df = pd.read_csv(input_csv)
    logger.info("  Loaded %d rows", len(df))

    if apply_consolidation:
        before = len(df[df["canonical_variable"] == "Temperature"])
        df = _consolidate_temperature(df)
        after = len(df[df["canonical_variable"] == "Temperature"])
        logger.info(
            "  Temperature consolidation: %d rows → %d rows  (dropped %d duplicates)",
            before, after, before - after,
        )

    # Write back the cleaned CSV so downstream scripts see the same data
    df.to_csv(input_csv, index=False)
    logger.info("  Cleaned CSV written back to %s", input_csv)

    # --- Build Excel ---
    logger.info("Writing Excel to %s", output_xlsx)
    with pd.ExcelWriter(output_xlsx, engine="openpyxl") as writer:
        # Write a blank placeholder so openpyxl creates the workbook
        pd.DataFrame().to_excel(writer, sheet_name="Matrix", index=False)

    wb = load_workbook(output_xlsx)

    # ── Sheet 1: Matrix ───────────────────────────────────────────────────
    ws_matrix = wb["Matrix"]
    _apply_matrix_sheet(ws_matrix, df)

    # ── Sheet 2: Legend ───────────────────────────────────────────────────
    ws_legend = wb.create_sheet("Legend")
    _apply_legend_sheet(ws_legend)

    # ── Sheet 3: Stats ────────────────────────────────────────────────────
    ws_stats = wb.create_sheet("Stats")
    _apply_stats_sheet(ws_stats, df)

    wb.save(output_xlsx)
    logger.info(
        "  Excel written: %d matrix rows across %d canonical variables",
        len(df),
        df["canonical_variable"].nunique(),
    )

    return df


# ---------------------------------------------------------------------------
# CLI entry-point
# ---------------------------------------------------------------------------

def main() -> None:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(name)s – %(message)s",
    )
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--no-consolidate",
        action="store_true",
        help="Skip Temperature consolidation and export the CSV as-is.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=COMBINATION_MATRIX_XLSX,
        help=f"Path to write the Excel file (default: {COMBINATION_MATRIX_XLSX})",
    )
    args = parser.parse_args()

    df = consolidate_and_export(
        output_xlsx=args.output,
        apply_consolidation=not args.no_consolidate,
    )

    print(f"\nFinal matrix: {len(df)} rows")
    summary = (
        df.groupby("canonical_variable")["subcategory"]
        .count()
        .reset_index()
        .rename(columns={"subcategory": "rows"})
    )
    print(summary.to_string(index=False))


if __name__ == "__main__":
    main()
