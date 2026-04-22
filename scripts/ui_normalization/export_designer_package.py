"""
Standalone script: generates the designer curation package in for_designer/.

Outputs:
  for_designer/Possibility_Flow.xlsx  - 3 sheets: Selection Flow, Combo Units, Connectors
  for_designer/EAP_Logic_Map.xlsx     - 2 sheets: Per EAP Logic Map, Activation Pattern Summary

Run:
  python -m scripts.ui_normalization.export_designer_package
"""

from __future__ import annotations

import json
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
ROOT = Path(__file__).resolve().parents[2]
UI_OUT = ROOT / "ui_normalization_output"
OUT_DIR = ROOT / "for_designer"

COMBO_CSV     = UI_OUT / "combination_matrix_enriched.csv"
TAXONOMY_CSV  = UI_OUT / "taxonomy_match_results_v2.csv"
CONNECTOR_JSON = UI_OUT / "connector_map.json"
SCHEMA_JSON   = UI_OUT / "ui_schema_openai_filtered.json"

# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")   # dark blue
ALT_FILL    = PatternFill("solid", fgColor="D6E4F0")   # light blue
WHITE_FILL  = PatternFill("solid", fgColor="FFFFFF")

HEADER_FONT  = Font(bold=True, color="FFFFFF", size=10)
BODY_FONT    = Font(size=10)
WRAP_ALIGN   = Alignment(wrap_text=True, vertical="top")
TOP_ALIGN    = Alignment(vertical="top")

THIN = Side(style="thin", color="CCCCCC")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Column widths (character units)
_COL_WIDTHS: dict[str, int] = {
    "Forecast Variable":              22,
    "Subcategory":                    28,
    "Unit":                           16,
    "Threshold Operator(s)":          22,
    "Accumulation Window?":           22,
    "Accumulation Window Unit(s)":    26,
    "Observed Accumulation Values Example": 30,
    "Persistence?":                          16,
    "Persistence Unit(s)":                   22,
    "Persistence Values Example":            28,
    "Probability Field?":                    18,
    "Observed Probability Range Example":    28,
    "Lead Time Field?":                      16,
    "Lead Time Unit(s)":                     18,
    "Observed Lead Time Range Example":      26,
    "Geographic Scope Required?":     24,
    "Geographic Scope Types":         36,
    "Geographic Label Examples":      38,
    "Applicable Hazards":             28,
    "Example Statement":              62,
    "Complete Flow":                  72,
    # Combo Units sheet
    "Accumulation Window Unit(s)_c":  26,
    "Combo Example":                  24,
    "Persistence Unit(s)_c":          22,
    "Persistence Example":            24,
    "What the Secondary Field Means": 34,
    "When to Show It":                36,
    # Connectors sheet
    "Connector (raw)":                18,
    "Plain-English Label":            26,
    "Recommended UI Label":           28,
    "Applies Between":                32,
    "Example Wording":                54,
    "Design Implication":             48,
    # EAP Logic Map sheet
    "EAP Document":                   40,
    "Hazard":                         16,
    "Activation Type":                22,
    "Phase Map":                      34,
    "Trigger Connectors":             24,
    "Inter-Phase Connector":          24,
    "Stop Mechanism?":                18,
    "Stop Connector":                 16,
    "Notes":                          54,
    "Activation Pattern":             28,
    "Count of EAPs":                  16,
    "Example EAP":                    40,
    "What it means for the UI":       52,
}


def _apply_header(ws, headers: list[str], row: int = 1) -> None:
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = WRAP_ALIGN
        cell.border = THIN_BORDER
        w = _COL_WIDTHS.get(h, max(len(h) + 4, 14))
        ws.column_dimensions[get_column_letter(col)].width = w


def _apply_body_row(ws, row: int, values: list, wrap_cols: set[int] | None = None) -> None:
    fill = ALT_FILL if row % 2 == 0 else WHITE_FILL
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val if val is not None else "")
        cell.font = BODY_FONT
        cell.fill = fill
        cell.border = THIN_BORDER
        if wrap_cols and col in wrap_cols:
            cell.alignment = WRAP_ALIGN
        else:
            cell.alignment = TOP_ALIGN


def _finalise_sheet(ws, n_cols: int) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


# ---------------------------------------------------------------------------
# Build taxonomy enrichment lookup tables
# ---------------------------------------------------------------------------

def _pipe(values: list[str]) -> str:
    """Join non-empty unique values with ' | '."""
    seen: list[str] = []
    for v in values:
        v = str(v).strip()
        if v and v.lower() not in ("nan", "none", "") and v not in seen:
            seen.append(v)
    return " | ".join(seen)


def _range_str(series: pd.Series) -> str:
    """Return 'min - max' for a numeric series, or '' if empty."""
    s = pd.to_numeric(series, errors="coerce").dropna()
    if s.empty:
        return ""
    lo, hi = s.min(), s.max()
    if lo == hi:
        return str(lo)
    return f"{lo:g} - {hi:g}"


def build_taxonomy_enrichment(df_tax: pd.DataFrame) -> dict:
    """
    Build per (matched_canonical, matched_subcategory, matched_unit) lookup dicts.
    Keys are (canonical, subcategory, unit) tuples; also a (canonical, subcategory) variant
    for fields that don't depend on unit.
    """
    # Keep only successfully parsed rows
    valid = df_tax[
        df_tax["llm_error"].isna() &
        df_tax["matched_canonical"].notna() &
        df_tax["matched_subcategory"].notna()
    ].copy()

    # Normalise key columns
    valid["_can"]  = valid["matched_canonical"].str.strip()
    valid["_sub"]  = valid["matched_subcategory"].str.strip()
    valid["_unit"] = valid["matched_unit"].fillna("").str.strip()

    # -----------------------------------------------------------------------
    # Aggregations grouped by (canonical, subcategory, unit)
    # -----------------------------------------------------------------------
    ops_by_unit: dict[tuple, str] = {}
    acc_vals_by_unit: dict[tuple, str] = {}

    for key, grp in valid.groupby(["_can", "_sub", "_unit"]):
        ops = sorted(grp["threshold_operator"].dropna().unique().tolist())
        ops_by_unit[key] = " | ".join(ops)

        acc_vals = sorted(
            {f"{v:g} {u}" for v, u in zip(
                pd.to_numeric(grp["accumulation_window_value"], errors="coerce").dropna(),
                grp.loc[pd.to_numeric(grp["accumulation_window_value"], errors="coerce").notna(), "accumulation_window_unit"].fillna("")
            ) if str(u).strip()}
        )
        acc_vals_by_unit[key] = " | ".join(acc_vals)

    # -----------------------------------------------------------------------
    # Aggregations grouped by (canonical, subcategory) - unit-independent
    # -----------------------------------------------------------------------
    geo_types_by_sub: dict[tuple, str] = {}
    geo_labels_by_sub: dict[tuple, str] = {}
    prob_range_by_sub: dict[tuple, str] = {}
    lt_range_by_sub: dict[tuple, str] = {}
    pers_vals_by_sub: dict[tuple, str] = {}
    ops_by_sub: dict[tuple, str] = {}          # fallback when unit not in ops_by_unit

    for key2, grp2 in valid.groupby(["_can", "_sub"]):
        # Geographic scope types
        types = sorted(grp2["geographic_scope_type"].dropna().unique().tolist())
        geo_types_by_sub[key2] = " | ".join(types)

        # Up to 3 real label examples
        labels = (
            grp2["geographic_scope_label"]
            .dropna()
            .loc[grp2["geographic_scope_label"].str.strip() != ""]
            .unique()
            .tolist()
        )
        geo_labels_by_sub[key2] = " | ".join(labels[:3])

        # Probability range
        prob_range_by_sub[key2] = _range_str(grp2["probability_value"])

        # Lead time range (use lead_time_value; fall back to timeframe_unit for unit)
        lt_range_by_sub[key2] = _range_str(grp2["lead_time_value"])

        # Persistence values
        pers_vals = sorted(
            {f"{v:g} {u}" for v, u in zip(
                pd.to_numeric(grp2["persistence_value"], errors="coerce").dropna(),
                grp2.loc[pd.to_numeric(grp2["persistence_value"], errors="coerce").notna(), "persistence_unit"].fillna("")
            ) if str(u).strip()}
        )
        pers_vals_by_sub[key2] = " | ".join(pers_vals)

        # Operator fallback (all operators for this canonical+subcategory)
        all_ops = sorted(grp2["threshold_operator"].dropna().unique().tolist())
        ops_by_sub[key2] = " | ".join(all_ops)

    return {
        "ops_by_unit":       ops_by_unit,
        "ops_by_sub":        ops_by_sub,
        "acc_vals_by_unit":  acc_vals_by_unit,
        "geo_types_by_sub":  geo_types_by_sub,
        "geo_labels_by_sub": geo_labels_by_sub,
        "prob_range_by_sub": prob_range_by_sub,
        "lt_range_by_sub":   lt_range_by_sub,
        "pers_vals_by_sub":  pers_vals_by_sub,
    }


def _nstr(val) -> str:
    """Safe string conversion: NaN / None / 'nan' -> empty string."""
    if val is None:
        return ""
    s = str(val).strip()
    return "" if s.lower() in ("nan", "none", "nat", "<na>") else s


def _schema_ops(schema: dict, canonical: str, subcategory: str) -> str:
    """Fallback: get operators from schema JSON."""
    overrides = schema.get("canonical_to_operators", {})
    ops = overrides.get(canonical, [])
    if ops:
        return " | ".join(sorted(ops))
    return ""


# ---------------------------------------------------------------------------
# Helper: flags
# ---------------------------------------------------------------------------

def _flag(val: str) -> str:
    v = str(val).strip().lower()
    if v in ("yes", "true"):
        return "Yes"
    if v in ("optional", "conditional"):
        return "Optional"
    return "No"


def _geo_required(val: str) -> str:
    v = str(val).strip().lower()
    if v == "always":
        return "Required"
    if v == "optional":
        return "Optional"
    return "Not required"


def _acc_secondary_applies(row: pd.Series) -> bool:
    v = str(row.get("accumulation_window_applicable", "no")).strip().lower()
    return v in ("yes", "optional", "conditional")


def _pers_secondary_applies(row: pd.Series) -> bool:
    v = str(row.get("persistence_applicable", "no")).strip().lower()
    return v in ("yes", "optional", "conditional")


def _when_to_show_acc(row: pd.Series) -> str:
    v = str(row.get("accumulation_window_applicable", "no")).strip().lower()
    if v == "yes":
        return "Always show"
    if v in ("optional", "conditional"):
        return "Show when unit is a volume/amount (mm, cm, inches, %)"
    return "Do not show"


def _when_to_show_pers(row: pd.Series) -> str:
    v = str(row.get("persistence_applicable", "no")).strip().lower()
    if v == "yes":
        return "Always show"
    if v in ("optional", "conditional"):
        return "Show when a sustained-duration condition applies"
    return "Do not show"


def _combo_example_acc(unit: str, acc_units_str: str) -> str:
    if not acc_units_str or acc_units_str.strip().lower() in ("", "n/a"):
        return ""
    wu = acc_units_str.split("|")[0].strip()
    if unit and unit.lower() not in ("-", ""):
        return f"e.g. 40 {unit} in 5 {wu}"
    return f"e.g. value in 5 {wu}"


def _combo_example_pers(unit: str, pers_units_str: str) -> str:
    if not pers_units_str or pers_units_str.strip().lower() in ("", "n/a"):
        return ""
    wu = pers_units_str.split("|")[0].strip()
    if unit and unit.lower() not in ("-", ""):
        return f"e.g. 3 {unit} for 12 {wu}"
    return f"e.g. value for 12 {wu}"


# ---------------------------------------------------------------------------
# Build Possibility_Flow.xlsx
# ---------------------------------------------------------------------------

CONNECTOR_VOCAB = [
    {
        "Connector (raw)": "AND",
        "Plain-English Label": "Both required (simultaneous)",
        "Recommended UI Label": "AND - all conditions must be met",
        "Applies Between": "Between trigger statements within the same activation phase",
        "Example Wording": "Pre-activation rainfall threshold AND activation river level threshold must both be met",
        "Design Implication": (
            "Show both trigger conditions in the UI with a visible AND connector. "
            "Activation is blocked if either condition is unmet."
        ),
    },
    {
        "Connector (raw)": "OR",
        "Plain-English Label": "Either condition is sufficient",
        "Recommended UI Label": "OR - any one condition is enough",
        "Applies Between": "Between trigger statements within the same activation phase",
        "Example Wording": "Trigger 1 (rainfall > 50 mm) OR Trigger 2 (river level > 3 m) is sufficient for activation",
        "Design Implication": (
            "Show both conditions but communicate that only one needs to be satisfied. "
            "Activation proceeds if any single condition is met."
        ),
    },
    {
        "Connector (raw)": "THEN",
        "Plain-English Label": "Phased / sequential (one follows the other)",
        "Recommended UI Label": "THEN - if this, then next step becomes active",
        "Applies Between": "Between a pre-activation phase and an activation phase (multi-stage EAPs)",
        "Example Wording": "Pre-activation seasonal forecast threshold is met THEN activation observational threshold becomes relevant",
        "Design Implication": (
            "Reflects a two-step process. The second step should only become visible or active "
            "after the first is confirmed. Consider a stepper or wizard pattern."
        ),
    },
    {
        "Connector (raw)": "ENABLES",
        "Plain-English Label": "Pre-condition unlocks the next phase",
        "Recommended UI Label": "ENABLES - pre-activation unlocks activation",
        "Applies Between": "Inter-phase: between the pre-activation phase and the activation phase",
        "Example Wording": "Pre-activation phase ENABLES the activation phase once the readiness threshold is crossed",
        "Design Implication": (
            "The activation phase should appear locked or inactive until the pre-activation "
            "condition is confirmed. Show a visual state change (e.g. greyed -> active)."
        ),
    },
    {
        "Connector (raw)": "PRECEDES",
        "Plain-English Label": "Informs but does not strictly gate the next phase",
        "Recommended UI Label": "PRECEDES - early signal informs but does not block",
        "Applies Between": "Inter-phase: a leading indicator phase that informs but does not gate activation",
        "Example Wording": "El Nino forecast PRECEDES and informs the main precipitation-based activation trigger",
        "Design Implication": (
            "Show the precursor signal as contextual information alongside the main activation "
            "condition. It does not block the workflow - it provides context."
        ),
    },
    {
        "Connector (raw)": "OPTIONAL_PRECURSOR",
        "Plain-English Label": "Optional early signal (may or may not be present)",
        "Recommended UI Label": "OPTIONAL PRECURSOR - advisory signal, not a gate",
        "Applies Between": "Inter-phase: optional pre-activation signal before the main activation phase",
        "Example Wording": "Optional pre-activation wind threshold may be monitored, with main activation based on population impact",
        "Design Implication": (
            "Display as a collapsible or secondary panel. It should not be required to complete "
            "the main activation trigger workflow."
        ),
    },
    {
        "Connector (raw)": "CANCELS",
        "Plain-English Label": "Stop / deactivation - ends the current response",
        "Recommended UI Label": "CANCELS - stop condition ends the response",
        "Applies Between": "A stop/deactivation condition that terminates an active EAP response",
        "Example Wording": "If river level drops below 2 m for 48 consecutive hours, CANCEL the active flood response",
        "Design Implication": (
            "Show as a dedicated deactivation condition separate from the activation flow. "
            "Consider a red or warning-coloured indicator to distinguish it from activation conditions."
        ),
    },
]


def build_possibility_flow(
    df_matrix: pd.DataFrame,
    enrich: dict,
    schema: dict,
) -> Workbook:
    wb = Workbook()

    # ------------------------------------------------------------------
    # Explode pipe-separated units: one row per unit
    # ------------------------------------------------------------------
    exploded_rows: list[dict] = []
    for _, row in df_matrix.iterrows():
        raw_units = str(row.get("primary_unit", "")).strip()
        units = [u.strip() for u in raw_units.split("|") if u.strip()] or ["-"]
        for unit in units:
            exploded_rows.append({**row.to_dict(), "_unit_exploded": unit})
    df_exp = pd.DataFrame(exploded_rows)

    # ------------------------------------------------------------------
    # Sheet 1 - Selection Flow
    # ------------------------------------------------------------------
    ws1 = wb.active
    ws1.title = "Selection Flow"

    headers1 = [
        "Forecast Variable",
        "Subcategory",
        "Unit",
        "Threshold Operator(s)",
        "Accumulation Window?",
        "Accumulation Window Unit(s)",
        "Observed Accumulation Values Example",
        "Persistence?",
        "Persistence Unit(s)",
        "Persistence Values Example",
        "Probability Field?",
        "Observed Probability Range Example",
        "Lead Time Field?",
        "Lead Time Unit(s)",
        "Observed Lead Time Range Example",
        "Geographic Scope Required?",
        "Geographic Scope Types",
        "Geographic Label Examples",
        "Applicable Hazards",
        "Example Statement",
        "Complete Flow",
    ]
    _apply_header(ws1, headers1)
    ws1.row_dimensions[1].height = 32

    WRAP_COLS_S1 = {18, 19, 20, 21}   # Geographic Label Examples, Hazards, Example Statement, Complete Flow

    data_row = 2
    for _, row in df_exp.iterrows():
        can  = str(row.get("canonical_variable", "")).strip()
        sub  = str(row.get("subcategory", "")).strip()
        unit = str(row.get("_unit_exploded", "")).strip()
        key3 = (can, sub, unit)
        key2 = (can, sub)

        # Operators: taxonomy first, schema fallback
        ops = enrich["ops_by_unit"].get(key3, "") or enrich["ops_by_sub"].get(key2, "") or _schema_ops(schema, can, sub)

        # Accumulation window
        acc_flag      = _flag(row.get("accumulation_window_applicable", "no"))
        raw_acc_units = _nstr(row.get("accumulation_window_units", ""))
        acc_units     = raw_acc_units.replace("|", " | ") if raw_acc_units and raw_acc_units.lower() not in ("n/a",) else "-"
        acc_vals      = enrich["acc_vals_by_unit"].get(key3, "")

        # Persistence
        pers_flag      = _flag(row.get("persistence_applicable", "no"))
        raw_pers_units = _nstr(row.get("persistence_units", ""))
        pers_units     = raw_pers_units.replace("|", " | ") if raw_pers_units and raw_pers_units.lower() not in ("n/a",) else "-"
        pers_vals      = enrich["pers_vals_by_sub"].get(key2, "")

        # Probability
        prob_flag  = _flag(row.get("probability_applicable", "no"))
        prob_range = enrich["prob_range_by_sub"].get(key2, "")
        prob_range_display = f"{prob_range}%" if prob_range else ""

        # Lead time
        lt_flag  = _flag(row.get("lead_time_applicable", "no"))
        raw_lt   = _nstr(row.get("lead_time_units", ""))
        lt_units = raw_lt.replace("|", " | ") if raw_lt and raw_lt.lower() not in ("n/a",) else "-"
        lt_range = enrich["lt_range_by_sub"].get(key2, "")

        # Geographic scope
        geo_req    = _geo_required(_nstr(row.get("geographic_scope_required", "optional")) or "optional")
        geo_types  = enrich["geo_types_by_sub"].get(key2, "")
        geo_labels = enrich["geo_labels_by_sub"].get(key2, "")

        hazards = _nstr(row.get("hazard_types", "")).replace("|", ", ")
        example = _nstr(row.get("example_statement", ""))

        # Complete Flow column - full decision path in one readable string
        flow_parts = [f"{can} > {sub} > {unit}"]
        if ops:
            flow_parts.append(f"Operator: {ops}")
        if acc_flag == "Yes" or (acc_flag == "Optional" and acc_units != "-"):
            acc_part = f"Acc Window: {acc_units}"
            if acc_vals:
                acc_part += f" (observed: {acc_vals})"
            flow_parts.append(acc_part)
        if pers_flag == "Yes" or (pers_flag == "Optional" and pers_units != "-"):
            pers_part = f"Persistence: {pers_units}"
            if pers_vals:
                pers_part += f" (observed: {pers_vals})"
            flow_parts.append(pers_part)
        if prob_flag != "No":
            prob_part = f"Probability: {prob_flag}"
            if prob_range_display:
                prob_part += f" (range: {prob_range_display})"
            flow_parts.append(prob_part)
        if lt_flag != "No":
            lt_part = f"Lead Time: {lt_units}"
            if lt_range:
                lt_part += f" (observed: {lt_range})"
            flow_parts.append(lt_part)
        if geo_types:
            flow_parts.append(f"Geo Scope [{geo_req}]: {geo_types}")
        else:
            flow_parts.append(f"Geo Scope: {geo_req}")
        complete_flow = " > ".join(flow_parts)

        _apply_body_row(ws1, data_row, [
            can, sub, unit,
            ops,
            acc_flag, acc_units, acc_vals,
            pers_flag, pers_units, pers_vals,
            prob_flag, prob_range_display,
            lt_flag, lt_units, lt_range,
            geo_req, geo_types, geo_labels,
            hazards, example,
            complete_flow,
        ], wrap_cols=WRAP_COLS_S1)
        ws1.row_dimensions[data_row].height = 42
        data_row += 1

    _finalise_sheet(ws1, len(headers1))

    # ------------------------------------------------------------------
    # Sheet 2 - Combo Units
    # ------------------------------------------------------------------
    ws2 = wb.create_sheet("Combo Units")

    headers2 = [
        "Forecast Variable",
        "Subcategory",
        "Unit",
        "Accumulation Window?",
        "Accumulation Window Unit(s)",
        "Combo Example",
        "Persistence?",
        "Persistence Unit(s)",
        "Persistence Example",
        "What the Secondary Field Means",
        "When to Show It",
    ]
    _apply_header(ws2, headers2)
    ws2.row_dimensions[1].height = 32

    data_row = 2
    seen_combos: set[tuple] = set()
    for _, row in df_exp.iterrows():
        can  = str(row.get("canonical_variable", "")).strip()
        sub  = str(row.get("subcategory", "")).strip()
        unit = str(row.get("_unit_exploded", "")).strip()

        has_acc  = _acc_secondary_applies(row)
        has_pers = _pers_secondary_applies(row)
        if not has_acc and not has_pers:
            continue

        combo_key = (can, sub, unit)
        if combo_key in seen_combos:
            continue
        seen_combos.add(combo_key)

        raw_acc_units  = _nstr(row.get("accumulation_window_units", ""))
        acc_units_str  = raw_acc_units if raw_acc_units and raw_acc_units.lower() not in ("n/a",) else "-"
        raw_pers_units = _nstr(row.get("persistence_units", ""))
        pers_units_str = raw_pers_units if raw_pers_units and raw_pers_units.lower() not in ("n/a",) else "-"

        combo_ex   = _combo_example_acc(unit, acc_units_str) if has_acc else "-"
        pers_ex    = _combo_example_pers(unit, pers_units_str) if has_pers else "-"

        meaning_parts = []
        if has_acc:
            meaning_parts.append(
                "Accumulation window - the time span over which the threshold amount accumulates "
                f"(e.g. '40 {unit} in 5 {acc_units_str.split('|')[0].strip()}')"
            )
        if has_pers:
            meaning_parts.append(
                "Persistence window - the duration the condition must be sustained continuously "
                f"(e.g. '3 {unit} for 12 {pers_units_str.split('|')[0].strip()}')"
            )
        meaning = "; ".join(meaning_parts)

        when_parts = []
        if has_acc:
            when_parts.append(_when_to_show_acc(row))
        if has_pers:
            when_parts.append(_when_to_show_pers(row))
        when_show = "; ".join(when_parts)

        _apply_body_row(ws2, data_row, [
            can, sub, unit,
            _flag(row.get("accumulation_window_applicable", "no")),
            acc_units_str.replace("|", " | "),
            combo_ex,
            _flag(row.get("persistence_applicable", "no")),
            pers_units_str.replace("|", " | "),
            pers_ex,
            meaning,
            when_show,
        ], wrap_cols={10, 11})
        ws2.row_dimensions[data_row].height = 58
        data_row += 1

    _finalise_sheet(ws2, len(headers2))

    # ------------------------------------------------------------------
    # Sheet 3 - Connectors
    # ------------------------------------------------------------------
    ws3 = wb.create_sheet("Connectors")

    headers3 = [
        "Connector (raw)",
        "Plain-English Label",
        "Recommended UI Label",
        "Applies Between",
        "Example Wording",
        "Design Implication",
    ]
    _apply_header(ws3, headers3)
    ws3.row_dimensions[1].height = 32

    for i, entry in enumerate(CONNECTOR_VOCAB, start=2):
        _apply_body_row(ws3, i, [
            entry["Connector (raw)"],
            entry["Plain-English Label"],
            entry["Recommended UI Label"],
            entry["Applies Between"],
            entry["Example Wording"],
            entry["Design Implication"],
        ], wrap_cols={4, 5, 6})
        ws3.row_dimensions[i].height = 72

    _finalise_sheet(ws3, len(headers3))

    # ------------------------------------------------------------------
    # Sheet 4 - Column Descriptions
    # ------------------------------------------------------------------
    ws4 = wb.create_sheet("Column Descriptions")
    ws4.sheet_properties.tabColor = "70AD47"  # green tab
    _apply_header(ws4, ["Sheet", "Column Name", "What it means", "Source"])
    ws4.column_dimensions["A"].width = 20
    ws4.column_dimensions["B"].width = 34
    ws4.column_dimensions["C"].width = 70
    ws4.column_dimensions["D"].width = 36
    ws4.row_dimensions[1].height = 28

    _PFLOW_DESCS = [
        # sheet, column, description, source
        ("Selection Flow", "Forecast Variable",
         "Top-level hazard/impact domain the trigger monitors (e.g. Precipitation, Hydrological Flow, Wind).",
         "combination_matrix_enriched.csv"),
        ("Selection Flow", "Subcategory",
         "Specific variable measured within the domain (e.g. Total rainfall, River flow, Seasonal total).",
         "combination_matrix_enriched.csv"),
        ("Selection Flow", "Unit",
         "Measurement unit for this row. Pipe-separated units from the matrix are exploded -- one row per unit.",
         "combination_matrix_enriched.csv"),
        ("Selection Flow", "Threshold Operator(s)",
         "Comparison operators seen in the real EAP corpus for this variable+subcategory+unit (e.g. >=, >, between). Falls back to schema JSON if not in corpus.",
         "taxonomy_match_results_v2.csv"),
        ("Selection Flow", "Accumulation Window?",
         "Whether an accumulation time window is part of this trigger: Yes = always required; Optional = depends on unit; No = never used.",
         "combination_matrix_enriched.csv"),
        ("Selection Flow", "Accumulation Window Unit(s)",
         "Time units supported for the accumulation window (e.g. hours, days, weeks). Pipe-separated.",
         "combination_matrix_enriched.csv"),
        ("Selection Flow", "Observed Accumulation Values Example",
         "Real accumulation window values found in EAP documents (e.g. 24 hours, 72 hours, 7 days). Use as examples to inform UI defaults.",
         "taxonomy_match_results_v2.csv"),
        ("Selection Flow", "Persistence?",
         "Whether the threshold must be sustained for a duration: Yes = always required; Optional = sometimes; No = not used.",
         "combination_matrix_enriched.csv"),
        ("Selection Flow", "Persistence Unit(s)",
         "Time units for the persistence window (e.g. days). Pipe-separated.",
         "combination_matrix_enriched.csv"),
        ("Selection Flow", "Persistence Values Example",
         "Real persistence durations found in EAP documents (e.g. 3 days, 4 days). Use as examples to inform UI defaults.",
         "taxonomy_match_results_v2.csv"),
        ("Selection Flow", "Probability Field?",
         "Whether a probability qualifier applies: Yes = always present; Optional = sometimes used; No = not relevant.",
         "combination_matrix_enriched.csv"),
        ("Selection Flow", "Observed Probability Range Example",
         "Min-max range of probability values seen in the EAP corpus for this variable (e.g. 40% - 80%). Expressed as percentages.",
         "taxonomy_match_results_v2.csv"),
        ("Selection Flow", "Lead Time Field?",
         "Whether a forecast lead time applies: Yes = always; Optional = sometimes; No = not used.",
         "combination_matrix_enriched.csv"),
        ("Selection Flow", "Lead Time Unit(s)",
         "Units for the forecast lead time (e.g. hours, days, months). Pipe-separated.",
         "combination_matrix_enriched.csv"),
        ("Selection Flow", "Observed Lead Time Range Example",
         "Min-max range of lead time values seen in the EAP corpus for this variable (in the units listed in Lead Time Unit(s)).",
         "taxonomy_match_results_v2.csv"),
        ("Selection Flow", "Geographic Scope Required?",
         "Whether geographic scoping is Required (always), Optional (sometimes), or Not required for this variable.",
         "combination_matrix_enriched.csv"),
        ("Selection Flow", "Geographic Scope Types",
         "Geographic scope categories observed in EAP documents (e.g. national, station_gauge, watershed_basin, administrative_unit). Pipe-separated.",
         "taxonomy_match_results_v2.csv"),
        ("Selection Flow", "Geographic Label Examples",
         "Up to 3 real place/scope label values extracted from EAP documents (e.g. Nowshera station, Sudd, flood-risk zones).",
         "taxonomy_match_results_v2.csv"),
        ("Selection Flow", "Applicable Hazards",
         "Hazard types this variable+subcategory+unit combination is used for (e.g. flood, drought, cholera). Comma-separated.",
         "combination_matrix_enriched.csv"),
        ("Selection Flow", "Example Statement",
         "A representative trigger statement text from the EAP corpus illustrating how this combination is used in practice.",
         "combination_matrix_enriched.csv"),
        ("Selection Flow", "Complete Flow",
         "Full end-to-end decision path in one readable string: Variable > Subcategory > Unit > Operator > Accumulation > Persistence > Probability > Lead Time > Geographic Scope. Mirrors the combination_path column in taxonomy_match_results_v2.csv but with all fields filled.",
         "Derived (all fields combined)"),
        ("Combo Units", "Forecast Variable", "Same as Selection Flow.", "combination_matrix_enriched.csv"),
        ("Combo Units", "Subcategory", "Same as Selection Flow.", "combination_matrix_enriched.csv"),
        ("Combo Units", "Unit", "Same as Selection Flow.", "combination_matrix_enriched.csv"),
        ("Combo Units", "Accumulation Window?",
         "Whether an accumulation time window applies. Only rows where Yes or Optional are shown in this sheet.",
         "combination_matrix_enriched.csv"),
        ("Combo Units", "Accumulation Window Unit(s)",
         "Supported time units for the accumulation window. Pipe-separated.",
         "combination_matrix_enriched.csv"),
        ("Combo Units", "Combo Example",
         "Worked example showing how a primary value and accumulation window combine: e.g. '40 mm in 5 days'.",
         "Derived"),
        ("Combo Units", "Persistence?",
         "Whether a persistence/duration condition applies.",
         "combination_matrix_enriched.csv"),
        ("Combo Units", "Persistence Unit(s)",
         "Supported time units for the persistence window.",
         "combination_matrix_enriched.csv"),
        ("Combo Units", "Persistence Example",
         "Worked example showing sustained-duration combination: e.g. '3 mm for 12 days'.",
         "Derived"),
        ("Combo Units", "What the Secondary Field Means",
         "Plain-English explanation of what the accumulation or persistence window captures for the designer.",
         "Derived"),
        ("Combo Units", "When to Show It",
         "UX guidance on when to render the secondary field: always, conditionally on unit, or never.",
         "Derived"),
        ("Connectors", "Connector (raw)",
         "The raw keyword used in EAP trigger logic (AND, OR, THEN, ENABLES, PRECEDES, OPTIONAL_PRECURSOR, CANCELS).",
         "connector_map.json"),
        ("Connectors", "Plain-English Label",
         "Human-readable description of what the connector means in the context of EAP activation.",
         "Derived"),
        ("Connectors", "Recommended UI Label",
         "Suggested label text for display in the designer tool when this connector is selected.",
         "Derived"),
        ("Connectors", "Applies Between",
         "Which parts of the trigger flow this connector bridges (within-phase vs inter-phase).",
         "Derived"),
        ("Connectors", "Example Wording",
         "Illustrative sentence showing how this connector appears in real EAP trigger language.",
         "Derived"),
        ("Connectors", "Design Implication",
         "UX design guidance on how to represent this connector in the interface.",
         "Derived"),
    ]

    desc_row = 2
    for sheet, col, meaning, source in _PFLOW_DESCS:
        fill = ALT_FILL if desc_row % 2 == 0 else WHITE_FILL
        for c, val in enumerate([sheet, col, meaning, source], 1):
            cell = ws4.cell(row=desc_row, column=c, value=val)
            cell.font = BODY_FONT
            cell.fill = fill
            cell.border = THIN_BORDER
            cell.alignment = WRAP_ALIGN
        ws4.row_dimensions[desc_row].height = 52
        desc_row += 1

    ws4.freeze_panes = "A2"

    return wb


# ---------------------------------------------------------------------------
# Build EAP_Logic_Map.xlsx
# ---------------------------------------------------------------------------

def _fmt_phase_map(phase_map: dict) -> str:
    return " -> ".join(
        f"{k}: {v.replace('_', ' ')}" for k, v in phase_map.items()
    )


def _fmt_connector_map(connector_map: dict) -> str:
    if not connector_map:
        return "None (single trigger)"
    return " | ".join(f"{k}: {v}" for k, v in connector_map.items())


ACTIVATION_LABELS = {
    "single-trigger":  "Single trigger",
    "dual-trigger":    "Dual trigger (two conditions)",
    "multi-stage":     "Multi-stage (sequential phases)",
}

PATTERN_UI_MEANING = {
    "single-trigger": (
        "One condition is sufficient. No connector needed in the UI. "
        "Simple threshold form with one trigger statement."
    ),
    "dual-trigger (AND)": (
        "Two conditions must both be met. Show both trigger conditions with a visible AND. "
        "Activation is blocked if either is unmet. Consider a checklist or step-based UI."
    ),
    "dual-trigger (OR)": (
        "Either condition alone is sufficient. Show both with an OR connector. "
        "Activation proceeds when any one is met."
    ),
    "multi-stage (THEN)": (
        "A pre-activation phase leads to an activation phase. The second step should "
        "only become actionable after the first is confirmed. Stepper/wizard UI recommended."
    ),
    "multi-stage (OPTIONAL_PRECURSOR)": (
        "An optional early signal precedes the main activation. "
        "Display as advisory context; it does not gate the workflow."
    ),
    "multi-stage (PRECEDES)": (
        "A precursor signal informs but does not strictly gate activation. "
        "Show as contextual background alongside the main trigger."
    ),
    "stop-enabled": (
        "A CANCELS connector ends the response. Needs a dedicated deactivation condition "
        "field clearly separated from the activation flow."
    ),
}


def _derive_pattern_key(eap: dict) -> str:
    atype = eap.get("activation_type", "")
    cmap  = eap.get("connector_map", {})
    inter = eap.get("inter_phase_connector") or ""
    connectors = set(cmap.values())
    if atype == "single-trigger":
        return "single-trigger"
    if atype == "dual-trigger":
        if "AND" in connectors:
            return "dual-trigger (AND)"
        if "OR" in connectors:
            return "dual-trigger (OR)"
        return "dual-trigger (AND)"
    if atype == "multi-stage":
        if inter == "OPTIONAL_PRECURSOR":
            return "multi-stage (OPTIONAL_PRECURSOR)"
        if inter == "PRECEDES":
            return "multi-stage (PRECEDES)"
        return "multi-stage (THEN)"
    return atype


def build_eap_logic_map(connector_data: list[dict]) -> Workbook:
    wb = Workbook()

    # Sheet 1 - Per EAP Logic Map
    ws1 = wb.active
    ws1.title = "Per EAP Logic Map"

    headers1 = [
        "EAP Document",
        "Hazard",
        "Activation Type",
        "Phase Map",
        "Trigger Connectors",
        "Inter-Phase Connector",
        "Stop Mechanism?",
        "Stop Connector",
        "Notes",
    ]
    _apply_header(ws1, headers1)
    ws1.row_dimensions[1].height = 32

    for i, eap in enumerate(connector_data, start=2):
        atype_raw   = eap.get("activation_type", "")
        atype_label = ACTIVATION_LABELS.get(atype_raw, atype_raw)
        stop_present = "Yes" if eap.get("stop_mechanism_present") else "No"
        stop_conn   = eap.get("stop_connector") or "-"
        inter       = eap.get("inter_phase_connector") or "-"

        _apply_body_row(ws1, i, [
            eap.get("document_name", ""),
            eap.get("hazard_type", ""),
            atype_label,
            _fmt_phase_map(eap.get("phase_map", {})),
            _fmt_connector_map(eap.get("connector_map", {})),
            inter,
            stop_present,
            stop_conn,
            eap.get("notes", ""),
        ], wrap_cols={4, 5, 9})
        ws1.row_dimensions[i].height = 55

    _finalise_sheet(ws1, len(headers1))

    # Sheet 2 - Activation Pattern Summary
    ws2 = wb.create_sheet("Activation Pattern Summary")

    headers2 = [
        "Activation Pattern",
        "Count of EAPs",
        "Example EAP",
        "What it means for the UI",
    ]
    _apply_header(ws2, headers2)
    ws2.row_dimensions[1].height = 32

    from collections import Counter, defaultdict
    pattern_counts: Counter = Counter()
    pattern_examples: defaultdict[str, str] = defaultdict(str)
    stop_eaps: list[str] = []

    for eap in connector_data:
        pk = _derive_pattern_key(eap)
        pattern_counts[pk] += 1
        if not pattern_examples[pk]:
            pattern_examples[pk] = eap.get("document_name", "")
        if eap.get("stop_mechanism_present"):
            stop_eaps.append(eap.get("document_name", ""))

    display_order = [
        "single-trigger",
        "dual-trigger (AND)",
        "dual-trigger (OR)",
        "multi-stage (THEN)",
        "multi-stage (PRECEDES)",
        "multi-stage (OPTIONAL_PRECURSOR)",
    ]

    data_row = 2
    for pk in display_order:
        if pk not in pattern_counts:
            continue
        _apply_body_row(ws2, data_row, [
            pk,
            pattern_counts[pk],
            pattern_examples[pk],
            PATTERN_UI_MEANING.get(pk, ""),
        ], wrap_cols={4})
        ws2.row_dimensions[data_row].height = 65
        data_row += 1

    _apply_body_row(ws2, data_row, [
        "stop-enabled (cross-cutting)",
        len(stop_eaps),
        stop_eaps[0] if stop_eaps else "-",
        PATTERN_UI_MEANING.get("stop-enabled", ""),
    ], wrap_cols={4})
    ws2.row_dimensions[data_row].height = 65

    _finalise_sheet(ws2, len(headers2))

    # ------------------------------------------------------------------
    # Sheet 3 - Column Descriptions (EAP Logic Map)
    # ------------------------------------------------------------------
    ws3 = wb.create_sheet("Column Descriptions")
    ws3.sheet_properties.tabColor = "70AD47"
    _apply_header(ws3, ["Sheet", "Column Name", "What it means", "Source"])
    ws3.column_dimensions["A"].width = 28
    ws3.column_dimensions["B"].width = 28
    ws3.column_dimensions["C"].width = 72
    ws3.column_dimensions["D"].width = 28
    ws3.row_dimensions[1].height = 28

    _EAP_DESCS = [
        ("Per EAP Logic Map", "EAP Document",
         "Full name of the EAP document as extracted from the source PDF.",
         "connector_map.json"),
        ("Per EAP Logic Map", "Hazard",
         "Primary hazard type this EAP responds to (e.g. Flood, Drought, Cholera).",
         "connector_map.json"),
        ("Per EAP Logic Map", "Activation Type",
         "Overall trigger structure: Single trigger (one condition), Dual trigger (two conditions), Multi-stage (sequential phases).",
         "connector_map.json"),
        ("Per EAP Logic Map", "Phase Map",
         "Ordered list of activation phases and their data type (e.g. pre_activation: forecast -> activation: observational). Arrow shows sequence.",
         "connector_map.json"),
        ("Per EAP Logic Map", "Trigger Connectors",
         "Connectors between individual trigger statements within a phase (AND = both required; OR = either sufficient).",
         "connector_map.json"),
        ("Per EAP Logic Map", "Inter-Phase Connector",
         "Connector linking the pre-activation phase to the activation phase (THEN, PRECEDES, OPTIONAL_PRECURSOR). '-' if single-phase.",
         "connector_map.json"),
        ("Per EAP Logic Map", "Stop Mechanism?",
         "Whether the EAP includes a deactivation/stop condition (CANCELS connector). Yes = a stop condition is defined.",
         "connector_map.json"),
        ("Per EAP Logic Map", "Stop Connector",
         "The connector keyword used for the stop/deactivation condition (typically CANCELS). '-' if no stop mechanism.",
         "connector_map.json"),
        ("Per EAP Logic Map", "Notes",
         "Additional context or caveats about the activation logic extracted from the EAP document.",
         "connector_map.json"),
        ("Activation Pattern Summary", "Activation Pattern",
         "Pattern category: single-trigger, dual-trigger (AND), dual-trigger (OR), multi-stage (THEN/PRECEDES/OPTIONAL_PRECURSOR), or stop-enabled.",
         "Derived from connector_map.json"),
        ("Activation Pattern Summary", "Count of EAPs",
         "Number of EAP documents in the corpus that use this activation pattern.",
         "Derived from connector_map.json"),
        ("Activation Pattern Summary", "Example EAP",
         "One representative EAP document name for this pattern.",
         "Derived from connector_map.json"),
        ("Activation Pattern Summary", "What it means for the UI",
         "Design guidance on how to represent this activation pattern in the trigger-builder interface.",
         "Derived"),
    ]

    desc_row = 2
    for sheet, col, meaning, source in _EAP_DESCS:
        fill = ALT_FILL if desc_row % 2 == 0 else WHITE_FILL
        for c, val in enumerate([sheet, col, meaning, source], 1):
            cell = ws3.cell(row=desc_row, column=c, value=val)
            cell.font = BODY_FONT
            cell.fill = fill
            cell.border = THIN_BORDER
            cell.alignment = WRAP_ALIGN
        ws3.row_dimensions[desc_row].height = 52
        desc_row += 1

    ws3.freeze_panes = "A2"

    return wb


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    OUT_DIR.mkdir(exist_ok=True)

    df_matrix  = pd.read_csv(COMBO_CSV)
    df_taxonomy = pd.read_csv(TAXONOMY_CSV, low_memory=False)

    with open(SCHEMA_JSON, encoding="utf-8") as f:
        schema = json.load(f)
    with open(CONNECTOR_JSON, encoding="utf-8") as f:
        connector_data = json.load(f)

    print("Building taxonomy enrichment lookups...")
    enrich = build_taxonomy_enrichment(df_taxonomy)

    wb1 = build_possibility_flow(df_matrix, enrich, schema)
    out1 = OUT_DIR / "Possibility_Flow.xlsx"
    wb1.save(out1)
    print(f"[OK] {out1}")

    wb2 = build_eap_logic_map(connector_data)
    out2 = OUT_DIR / "EAP_Logic_Map.xlsx"
    wb2.save(out2)
    print(f"[OK] {out2}")


if __name__ == "__main__":
    main()
